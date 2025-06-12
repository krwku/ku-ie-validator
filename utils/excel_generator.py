import tempfile
import os
from pathlib import Path

def classify_course(course_code, course_name=""):
    """
    Classify course into appropriate category based on course code.
    Returns: (category, subcategory, is_identified)
    """
    code = course_code.upper()
    
    # IE Core Courses (206xxx, 204111, 208xxx, 213xxx, 403xxx, 417xxx, 420xxx, 205xxx)
    if any(code.startswith(prefix) for prefix in ["01206", "01204111", "01208", "01213", "01403", "01417", "01420", "01205"]):
        return ("ie_core", "core", True)
    
    # Gen-Ed Classification
    if code.startswith("01175") or code.startswith("01101102") or code.startswith("01173151") or code.startswith("01999"):
        if any(keyword in course_name.lower() for keyword in ["health", "sport", "wellness", "aids", "food", "track", "tennis", "swim"]):
            return ("gen_ed", "wellness", True)
        elif any(keyword in course_name.lower() for keyword in ["art", "music", "design", "culture", "aesthetic"]):
            return ("gen_ed", "aesthetics", True)
        else:
            return ("gen_ed", "wellness", True)  # Default wellness for 999 courses
    
    if any(code.startswith(prefix) for prefix in ["01131", "01132", "01005", "01200", "01418102"]):
        return ("gen_ed", "entrepreneurship", True)
    
    if any(code.startswith(prefix) for prefix in ["01361", "01355", "01354", "01356", "01357", "01358", "01362", "01367", "01395", "01398", "01399", "01371", "01418104", "01418111"]):
        return ("gen_ed", "language_communication", True)
    
    if any(code.startswith(prefix) for prefix in ["01001", "01015", "01174", "01350", "01387104", "01390", "01301", "01455", "01460"]) or "999111" in code:
        return ("gen_ed", "thai_citizen_global", True)
    
    if any(code.startswith(prefix) for prefix in ["01007", "01240", "01373", "01376", "01387102", "01420201"]):
        return ("gen_ed", "aesthetics", True)
    
    # Technical Electives (advanced 206xxx courses not in core)
    if code.startswith("01206") and any(advanced in code for advanced in ["411", "412", "413", "421", "422", "423", "441", "442", "443", "461", "462", "463", "464", "490", "496", "498"]):
        return ("technical_electives", "technical", True)
    
    # Known Free Electives (courses that are intentionally free electives)
    # Add specific course codes that you know should be free electives
    known_free_electives = ["01206490", "01206499", "01206497"]  # Example: Coop, Project, Seminar
    if code in known_free_electives:
        return ("free_electives", "free", True)
    
    # Unidentified courses (everything else that doesn't match known patterns)
    return ("unidentified", "unknown", False)

def create_semester_based_worksheet(wb, student_info, semesters, validation_results):
    """Create semester-based worksheet similar to the provided image."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    ws = wb.create_sheet("Semester Based")
    
    # Define styles
    header_font = Font(bold=True, size=12)
    subheader_font = Font(bold=True, size=10)
    normal_font = Font(size=9)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Color fills
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    year_fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
    semester_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    core_fill = PatternFill(start_color="E6F3E6", end_color="E6F3E6", fill_type="solid")
    gen_ed_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    elective_fill = PatternFill(start_color="F2E6FF", end_color="F2E6FF", fill_type="solid")
    invalid_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
    unidentified_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    
    # Set column widths
    column_widths = {'A': 15, 'B': 35, 'C': 8, 'D': 8, 'E': 15, 'F': 35, 'G': 8, 'H': 8}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Header
    ws['A1'] = "KU INDUSTRIAL ENGINEERING - SEMESTER BASED PLAN"
    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws['A1'].fill = header_fill
    ws.merge_cells('A1:H1')
    ws['A1'].alignment = center_align
    
    # Student info
    ws['A3'] = f"Student: {student_info.get('name', '')} ({student_info.get('id', '')})"
    ws['A3'].font = subheader_font
    
    current_row = 5
    
    # Group semesters by year
    semesters_by_year = {}
    for semester in semesters:
        year = semester.get('year_int', 0)
        if year not in semesters_by_year:
            semesters_by_year[year] = []
        semesters_by_year[year].append(semester)
    
    # Sort years
    sorted_years = sorted([y for y in semesters_by_year.keys() if y > 0])
    
    for year_num, year in enumerate(sorted_years, 1):
        # Year header
        ws[f'A{current_row}'] = f"Year {year_num} ({year})"
        ws[f'A{current_row}'].font = header_font
        ws[f'A{current_row}'].fill = year_fill
        ws.merge_cells(f'A{current_row}:H{current_row}')
        ws[f'A{current_row}'].alignment = center_align
        current_row += 1
        
        year_semesters = semesters_by_year[year]
        year_semesters.sort(key=lambda s: s.get('semester_order', 0))
        
        # Semester headers
        col_positions = ['A', 'E']  # Two semesters per row
        semester_headers = []
        
        for i, semester in enumerate(year_semesters[:2]):  # Max 2 semesters per row
            col = col_positions[i]
            ws[f'{col}{current_row}'] = semester.get('semester', '')
            ws[f'{col}{current_row}'].font = subheader_font
            ws[f'{col}{current_row}'].fill = semester_fill
            ws[f'{col}{current_row}'].alignment = center_align
            ws[f'{col}{current_row}'].border = border
            
            # Course headers
            next_col = chr(ord(col) + 1)
            ws[f'{col}{current_row + 1}'] = "Course Code"
            ws[f'{next_col}{current_row + 1}'] = "Course Name"
            ws[f'{chr(ord(col) + 2)}{current_row + 1}'] = "Grade"
            ws[f'{chr(ord(col) + 3)}{current_row + 1}'] = "Credits"
            
            for j in range(4):
                cell = ws[f'{chr(ord(col) + j)}{current_row + 1}']
                cell.font = subheader_font
                cell.fill = semester_fill
                cell.alignment = center_align
                cell.border = border
        
        current_row += 2
        
        # Find maximum courses in any semester for this year
        max_courses = max(len(sem.get('courses', [])) for sem in year_semesters) if year_semesters else 0
        
        # Add courses
        for course_row in range(max_courses):
            for i, semester in enumerate(year_semesters[:2]):
                col = col_positions[i]
                courses = semester.get('courses', [])
                
                if course_row < len(courses):
                    course = courses[course_row]
                    course_code = course.get('code', '')
                    course_name = course.get('name', '')
                    grade = course.get('grade', '')
                    credits = course.get('credits', 0)
                    
                    # Check if course is valid
                    is_valid = True
                    for result in validation_results:
                        if (result.get('course_code') == course_code and 
                            not result.get('is_valid', True) and 
                            result.get('course_code') != 'CREDIT_LIMIT'):
                            is_valid = False
                            break
                    
                    # Classify course for coloring
                    category, subcategory, is_identified = classify_course(course_code, course_name)
                    
                    # Determine cell color
                    if not is_valid:
                        fill_color = invalid_fill
                    elif not is_identified:
                        fill_color = unidentified_fill
                    elif category == "ie_core":
                        fill_color = core_fill
                    elif category == "gen_ed":
                        fill_color = gen_ed_fill
                    else:
                        fill_color = elective_fill
                    
                    # Fill cells
                    ws[f'{col}{current_row + course_row}'] = course_code
                    ws[f'{chr(ord(col) + 1)}{current_row + course_row}'] = course_name
                    ws[f'{chr(ord(col) + 2)}{current_row + course_row}'] = grade
                    ws[f'{chr(ord(col) + 3)}{current_row + course_row}'] = credits
                    
                    for j in range(4):
                        cell = ws[f'{chr(ord(col) + j)}{current_row + course_row}']
                        cell.border = border
                        cell.fill = fill_color
                        cell.font = normal_font
                        if j == 1:  # Course name
                            cell.alignment = left_align
                        else:
                            cell.alignment = center_align
        
        current_row += max_courses + 2
    
    return ws

def create_requirements_based_worksheet(wb, student_info, semesters, validation_results):
    """Create requirements-based worksheet with improved semester identification."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    ws = wb.create_sheet("Requirements Based")
    
    # Define styles (same as before)
    header_font = Font(bold=True, size=11)
    subheader_font = Font(bold=True, size=10)
    normal_font = Font(size=9)
    small_font = Font(size=8)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Color fills
    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    gray_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    blue_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
    red_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
    orange_fill = PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid")
    
    # Set column widths
    column_widths = {
        'A': 12, 'B': 35, 'C': 8, 'D': 8, 'E': 15, 'F': 35, 'G': 8, 'H': 8,
        'I': 15, 'J': 35, 'K': 8, 'L': 8, 'M': 15, 'N': 35, 'O': 8, 'P': 8
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # HEADER SECTION
    ws['A1'] = "KU INDUSTRIAL ENGINEERING - REQUIREMENTS BASED PLAN"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:P1')
    ws['A1'].alignment = center_align
    ws['A1'].fill = blue_fill
    
    # Student Info
    ws['A3'] = "Student ID:"
    ws['B3'] = student_info.get('id', '')
    ws['B3'].fill = yellow_fill
    ws['E3'] = "Name:"
    ws['F3'] = student_info.get('name', '')
    ws['F3'].fill = yellow_fill
    
    # CLASSIFY ALL COURSES
    classified_courses = {
        "ie_core": [],
        "gen_ed": {
            "wellness": [],
            "entrepreneurship": [],
            "language_communication": [],
            "thai_citizen_global": [],
            "aesthetics": []
        },
        "technical_electives": [],
        "free_electives": [],
        "unidentified": []  # New category for unidentified courses
    }
    
    # Process all courses from all semesters
    for sem_idx, semester in enumerate(semesters):
        semester_name = semester.get("semester", f"Semester {sem_idx + 1}")
        year = semester.get("year_int", 0)
        semester_type = semester.get("semester_type", "")
        
        for course in semester.get("courses", []):
            course_code = course.get("code", "")
            course_name = course.get("name", "")
            grade = course.get("grade", "")
            credits = course.get("credits", 0)
            
            # Check if course is valid
            is_valid = True
            issue_reason = ""
            for result in validation_results:
                if (result.get('course_code') == course_code and 
                    not result.get('is_valid', True) and 
                    result.get('course_code') != 'CREDIT_LIMIT'):
                    is_valid = False
                    issue_reason = result.get('reason', '')
                    break
            
            # Classify course
            category, subcategory, is_identified = classify_course(course_code, course_name)
            
            course_info = {
                "code": course_code,
                "name": course_name,
                "grade": grade,
                "credits": credits,
                "semester": semester_name,
                "is_valid": is_valid,
                "issue": issue_reason,
                "year": year,
                "semester_type": semester_type,
                "is_identified": is_identified
            }
            
            # Place in appropriate category
            if not is_identified:
                classified_courses["unidentified"].append(course_info)
            elif category == "ie_core":
                classified_courses["ie_core"].append(course_info)
            elif category == "gen_ed":
                classified_courses["gen_ed"][subcategory].append(course_info)
            elif category == "technical_electives":
                classified_courses["technical_electives"].append(course_info)
            else:
                classified_courses["free_electives"].append(course_info)
    
    current_row = 6
    
    # IE CORE COURSES SECTION with better semester grouping
    ws[f'A{current_row}'] = "IE CORE COURSES"
    ws[f'A{current_row}'].font = header_font
    ws[f'A{current_row}'].fill = blue_fill
    ws.merge_cells(f'A{current_row}:P{current_row}')
    current_row += 1
    
    # Group IE courses by actual semesters, not just years
    ie_by_semester = {}
    for course in classified_courses["ie_core"]:
        semester_key = course["semester"]
        if semester_key not in ie_by_semester:
            ie_by_semester[semester_key] = []
        ie_by_semester[semester_key].append(course)
    
    # Display IE courses by semester
    for semester_name, courses in ie_by_semester.items():
        if courses:
            ws[f'A{current_row}'] = semester_name
            ws[f'A{current_row}'].font = subheader_font
            ws[f'A{current_row}'].fill = gray_fill
            current_row += 1
            
            # Headers
            headers = ["Code", "Course Name", "Grade", "Credits"]
            for i, header in enumerate(headers):
                cell = ws[f'{chr(65 + i)}{current_row}']
                cell.value = header
                cell.font = subheader_font
                cell.alignment = center_align
                cell.fill = gray_fill
                cell.border = border
            current_row += 1
            
            # Add courses
            for course in courses:
                ws[f'A{current_row}'] = course["code"]
                ws[f'B{current_row}'] = course["name"]
                ws[f'C{current_row}'] = course["grade"]
                ws[f'D{current_row}'] = course["credits"]
                
                # Apply styling
                for col in ['A', 'B', 'C', 'D']:
                    cell = ws[f'{col}{current_row}']
                    cell.border = border
                    cell.font = small_font
                    if col == 'B':
                        cell.alignment = left_align
                    else:
                        cell.alignment = center_align
                    
                    # Color coding
                    if not course["is_valid"]:
                        cell.fill = red_fill
                    elif course["grade"] not in ['F', 'W', 'N', '']:
                        cell.fill = green_fill
                
                current_row += 1
            current_row += 1
    
    # GEN-ED SECTIONS (same as before but with semester info)
    gen_ed_categories = [
        ("wellness", "WELLNESS (7 Credits Required)"),
        ("entrepreneurship", "ENTREPRENEURSHIP (5 Credits Required)"),
        ("language_communication", "LANGUAGE & COMMUNICATION (15 Credits Required)"),
        ("thai_citizen_global", "THAI CITIZEN & GLOBAL CITIZENSHIP (2 Credits Required)"),
        ("aesthetics", "AESTHETICS (3 Credits Required)")
    ]
    
    for category_key, category_name in gen_ed_categories:
        courses = classified_courses["gen_ed"][category_key]
        if courses:
            ws[f'A{current_row}'] = category_name
            ws[f'A{current_row}'].font = header_font
            ws[f'A{current_row}'].fill = gray_fill
            ws.merge_cells(f'A{current_row}:E{current_row}')
            current_row += 1
            
            for course in courses:
                ws[f'A{current_row}'] = course["code"]
                ws[f'B{current_row}'] = course["name"]
                ws[f'C{current_row}'] = course["grade"]
                ws[f'D{current_row}'] = course["credits"]
                ws[f'E{current_row}'] = course["semester"]
                
                # Apply styling and color coding
                for col in ['A', 'B', 'C', 'D', 'E']:
                    cell = ws[f'{col}{current_row}']
                    cell.border = border
                    cell.font = small_font
                    if col == 'B':
                        cell.alignment = left_align
                    else:
                        cell.alignment = center_align
                    
                    if not course["is_valid"]:
                        cell.fill = red_fill
                    elif course["grade"] not in ['F', 'W', 'N', '']:
                        cell.fill = green_fill
                
                current_row += 1
            current_row += 1
    
    # TECHNICAL ELECTIVES SECTION
    if classified_courses["technical_electives"]:
        ws[f'A{current_row}'] = "TECHNICAL ELECTIVES"
        ws[f'A{current_row}'].font = header_font
        ws[f'A{current_row}'].fill = blue_fill
        ws.merge_cells(f'A{current_row}:E{current_row}')
        current_row += 1
        
        for course in classified_courses["technical_electives"]:
            ws[f'A{current_row}'] = course["code"]
            ws[f'B{current_row}'] = course["name"]
            ws[f'C{current_row}'] = course["grade"]
            ws[f'D{current_row}'] = course["credits"]
            ws[f'E{current_row}'] = course["semester"]
            
            for col in ['A', 'B', 'C', 'D', 'E']:
                cell = ws[f'{col}{current_row}']
                cell.border = border
                cell.font = small_font
                if col == 'B':
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align
                
                if not course["is_valid"]:
                    cell.fill = red_fill
                elif course["grade"] not in ['F', 'W', 'N', '']:
                    cell.fill = green_fill
            
            current_row += 1
        current_row += 1
    
    # FREE ELECTIVES SECTION
    if classified_courses["free_electives"]:
        ws[f'A{current_row}'] = "FREE ELECTIVES"
        ws[f'A{current_row}'].font = header_font
        ws[f'A{current_row}'].fill = yellow_fill
        ws.merge_cells(f'A{current_row}:E{current_row}')
        current_row += 1
        
        for course in classified_courses["free_electives"]:
            ws[f'A{current_row}'] = course["code"]
            ws[f'B{current_row}'] = course["name"]
            ws[f'C{current_row}'] = course["grade"]
            ws[f'D{current_row}'] = course["credits"]
            ws[f'E{current_row}'] = course["semester"]
            
            for col in ['A', 'B', 'C', 'D', 'E']:
                cell = ws[f'{col}{current_row}']
                cell.border = border
                cell.font = small_font
                if col == 'B':
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align
                
                if not course["is_valid"]:
                    cell.fill = red_fill
                elif course["grade"] not in ['F', 'W', 'N', '']:
                    cell.fill = green_fill
            
            current_row += 1
        current_row += 1
    
    # UNIDENTIFIED COURSES SECTION - NEW!
    if classified_courses["unidentified"]:
        ws[f'A{current_row}'] = "⚠️ UNIDENTIFIED COURSES (PLEASE UPDATE DATABASE)"
        ws[f'A{current_row}'].font = header_font
        ws[f'A{current_row}'].fill = orange_fill
        ws.merge_cells(f'A{current_row}:E{current_row}')
        current_row += 1
        
        ws[f'A{current_row}'] = "These courses are not in the course database. Please verify and update the classification:"
        ws[f'A{current_row}'].font = Font(size=9, italic=True)
        ws.merge_cells(f'A{current_row}:E{current_row}')
        current_row += 1
        
        for course in classified_courses["unidentified"]:
            ws[f'A{current_row}'] = course["code"]
            ws[f'B{current_row}'] = course["name"]
            ws[f'C{current_row}'] = course["grade"]
            ws[f'D{current_row}'] = course["credits"]
            ws[f'E{current_row}'] = course["semester"]
            
            for col in ['A', 'B', 'C', 'D', 'E']:
                cell = ws[f'{col}{current_row}']
                cell.border = border
                cell.font = small_font
                cell.fill = orange_fill  # Always orange for unidentified
                if col == 'B':
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align
            
            current_row += 1
        current_row += 1
    
    # SUMMARY SECTION
    current_row += 1
    ws[f'A{current_row}'] = "CREDIT SUMMARY"
    ws[f'A{current_row}'].font = header_font
    ws[f'A{current_row}'].fill = blue_fill
    ws.merge_cells(f'A{current_row}:E{current_row}')
    current_row += 1
    
    # Calculate credits by category
    ie_credits = sum(c["credits"] for c in classified_courses["ie_core"] if c["grade"] not in ['F', 'W', 'N', ''])
    gen_ed_credits = sum(
        sum(c["credits"] for c in courses if c["grade"] not in ['F', 'W', 'N', ''])
        for courses in classified_courses["gen_ed"].values()
    )
    tech_credits = sum(c["credits"] for c in classified_courses["technical_electives"] if c["grade"] not in ['F', 'W', 'N', ''])
    free_credits = sum(c["credits"] for c in classified_courses["free_electives"] if c["grade"] not in ['F', 'W', 'N', ''])
    unidentified_credits = sum(c["credits"] for c in classified_courses["unidentified"] if c["grade"] not in ['F', 'W', 'N', ''])
    
    summary_data = [
        ("IE Core Courses:", ie_credits),
        ("Gen-Ed Courses:", gen_ed_credits),
        ("Technical Electives:", tech_credits),
        ("Free Electives:", free_credits),
        ("Unidentified Courses:", unidentified_credits),
        ("TOTAL CREDITS:", ie_credits + gen_ed_credits + tech_credits + free_credits + unidentified_credits)
    ]
    
    for i, (label, credits) in enumerate(summary_data):
        row = current_row + i
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = subheader_font
        ws[f'B{row}'] = credits
        ws[f'B{row}'].font = subheader_font
        
        if "Unidentified" in label:
            ws[f'A{row}'].fill = orange_fill
            ws[f'B{row}'].fill = orange_fill
        elif label.startswith("TOTAL"):
            ws[f'A{row}'].fill = blue_fill
            ws[f'B{row}'].fill = blue_fill
        else:
            ws[f'B{row}'].fill = green_fill if credits > 0 else yellow_fill
    
    return ws

def create_smart_registration_excel(student_info, semesters, validation_results):
    """
    Create a comprehensive Excel registration format with two worksheets.
    """
    try:
        from openpyxl import Workbook
        
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create both worksheets
        ws1 = create_semester_based_worksheet(wb, student_info, semesters, validation_results)
        ws2 = create_requirements_based_worksheet(wb, student_info, semesters, validation_results)
        
        # Make semester-based the active sheet
        wb.active = ws1
        
        # Save to bytes
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            wb.save(tmp_file.name)
            
            with open(tmp_file.name, 'rb') as f:
                excel_bytes = f.read()
            
            os.unlink(tmp_file.name)
            return excel_bytes
            
    except Exception as e:
        raise Exception(f"Error creating Excel file: {e}")

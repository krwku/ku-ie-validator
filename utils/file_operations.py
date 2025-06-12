#!/usr/bin/env python3
"""
File operations for saving and loading transcript data.
"""
import json
import logging
import os
from pathlib import Path

logger = logging.getLogger("file_operations")

def save_transcript(model, file_path):
    """
    Save transcript data to the specified file.
    
    Args:
        model: The transcript model to save
        file_path: Path to save the file
        
    Returns:
        True if save was successful, False otherwise
    """
    try:
        # Prepare data
        data = model.to_dict()
        
        # Write to file
        with open(file_path, 'w', encoding='utf-8') as file:
            json.dump(data, file, indent=2)
        
        logger.info(f"Saved transcript to {file_path}")
        return True
    
    except Exception as e:
        logger.error(f"Failed to save file: {e}")
        return False

def load_transcript(model, file_path):
    """
    Load transcript data from the specified file.
    
    Args:
        model: The transcript model to update
        file_path: Path to the file to load
        
    Returns:
        True if load was successful, False otherwise
    """
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            data = json.load(file)
        
        # Validate basic structure
        if not isinstance(data, dict) or "student_info" not in data or "semesters" not in data:
            logger.error(f"Invalid transcript data format: {file_path}")
            return False
        
        # Update model
        model.from_dict(data)
        model.set_changed(False)
        
        logger.info(f"Loaded transcript from {file_path}")
        return True
    
    except Exception as e:
        logger.error(f"Failed to load file: {e}")
        return False

def load_course_data(course_data_path):
    """
    Load course data from JSON file.
    
    Args:
        course_data_path: Path to course data JSON file
        
    Returns:
        Dictionary containing course data or empty dict on failure
    """
    try:
        if os.path.exists(course_data_path):
            with open(course_data_path, 'r', encoding='utf-8') as file:
                course_data = json.load(file)
                
            # Create a flattened dictionary of all courses for easy lookup
            all_courses = {}
            for course in course_data.get("industrial_engineering_courses", []):
                all_courses[course["code"]] = course
            
            logger.info(f"Loaded {len(all_courses)} courses from {course_data_path}")
            
            return {
                "raw_data": course_data, 
                "all_courses": all_courses
            }
        else:
            logger.warning(f"Course data file not found: {course_data_path}")
            return {}
    
    except Exception as e:
        logger.error(f"Error loading course data: {e}")
        return {}

def save_validation_report_excel(student_info, semesters, validation_results, file_path):
    """
    Save validation results to Excel file.
    
    Args:
        student_info: Dictionary with student information
        semesters: List of semester dictionaries
        validation_results: List of validation result dictionaries
        file_path: Path to save the Excel file
        
    Returns:
        True if save was successful, False otherwise
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from datetime import datetime
        
        def calculate_gpa(courses):
            """Calculate GPA using weighted average."""
            grade_points = {
                "A": 4.0, "B+": 3.5, "B": 3.0, "C+": 2.5, "C": 2.0, 
                "D+": 1.5, "D": 1.0, "F": 0.0
            }
            
            total_points = 0.0
            total_credits = 0
            
            for course in courses:
                grade = course.get("grade", "")
                credits = course.get("credits", 0)
                
                # Skip grades that don't contribute to GPA
                if grade not in grade_points:
                    continue
                
                total_points += grade_points[grade] * credits
                total_credits += credits
            
            return round(total_points / total_credits, 2) if total_credits > 0 else 0.0
        
        def calculate_cumulative_gpa(semesters, up_to_index, valid_only=False):
            """Calculate cumulative GPA up to specified semester."""
            all_courses = []
            for i in range(up_to_index + 1):
                semester_courses = semesters[i].get("courses", [])
                
                if valid_only:
                    # Filter to only valid courses
                    results_for_semester = [r for r in validation_results 
                                          if r.get("semester_index") == i and r.get("course_code") != "CREDIT_LIMIT"]
                    valid_course_codes = {r.get("course_code") for r in results_for_semester if r.get("is_valid", True)}
                    semester_courses = [c for c in semester_courses if c.get("code") in valid_course_codes]
                
                all_courses.extend(semester_courses)
            
            return calculate_gpa(all_courses)
        
        wb = Workbook()
        
        # Remove default sheet and create Summary sheet
        wb.remove(wb.active)
        summary_ws = wb.create_sheet("Summary")
        
        # Student Information section
        summary_ws['A1'] = "COURSE REGISTRATION VALIDATION REPORT"
        summary_ws['A1'].font = Font(bold=True, size=14)
        summary_ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        summary_ws['A4'] = "STUDENT INFORMATION"
        summary_ws['A4'].font = Font(bold=True)
        summary_ws['A5'] = "Student ID:"
        summary_ws['B5'] = student_info.get('id', 'Unknown')
        summary_ws['A6'] = "Name:"
        summary_ws['B6'] = student_info.get('name', 'Unknown')
        summary_ws['A7'] = "Field of Study:"
        summary_ws['B7'] = student_info.get('field_of_study', 'Unknown')
        summary_ws['A8'] = "Date of Admission:"
        summary_ws['B8'] = student_info.get('date_admission', 'Unknown')
        
        # Validation Summary
        invalid_count = len([r for r in validation_results if not r.get("is_valid", True)])
        summary_ws['A10'] = "VALIDATION SUMMARY"
        summary_ws['A10'].font = Font(bold=True)
        summary_ws['A11'] = "Semesters Analyzed:"
        summary_ws['B11'] = len(semesters)
        summary_ws['A12'] = "Registrations Checked:"
        summary_ws['B12'] = len(validation_results)
        summary_ws['A13'] = "Invalid Registrations:"
        summary_ws['B13'] = invalid_count
        
        # Semester Details with GPA calculations
        summary_ws['A15'] = "SEMESTER DETAILS"
        summary_ws['A15'].font = Font(bold=True)
        
        current_row = 17
        
        for i, semester in enumerate(semesters):
            # Semester header
            semester_name = semester.get("semester", f"Semester {i+1}")
            summary_ws.cell(row=current_row, column=1, value=semester_name).font = Font(bold=True)
            current_row += 1
            
            # Total credits
            total_credits = semester.get("total_credits", 0)
            summary_ws.cell(row=current_row, column=1, value=f"Total Credits: {total_credits}")
            current_row += 1
            
            # Calculate GPAs
            semester_courses = semester.get("courses", [])
            overall_sem_gpa = calculate_gpa(semester_courses)
            overall_cum_gpa = calculate_cumulative_gpa(semesters, i)
            
            # Overall GPA row
            summary_ws.cell(row=current_row, column=1, value=f"Overall - Semester GPA: {overall_sem_gpa}, Cumulative GPA: {overall_cum_gpa}")
            current_row += 1
            
            # Check if this semester has invalid courses
            semester_invalid_courses = [r for r in validation_results 
                                      if r.get("semester_index") == i and not r.get("is_valid", True) and r.get("course_code") != "CREDIT_LIMIT"]
            
            if semester_invalid_courses:
                # Calculate valid-only GPAs
                valid_results = [r for r in validation_results 
                               if r.get("semester_index") == i and r.get("course_code") != "CREDIT_LIMIT"]
                valid_course_codes = {r.get("course_code") for r in valid_results if r.get("is_valid", True)}
                valid_courses = [c for c in semester_courses if c.get("code") in valid_course_codes]
                
                valid_sem_gpa = calculate_gpa(valid_courses)
                valid_cum_gpa = calculate_cumulative_gpa(semesters, i, valid_only=True)
                
                summary_ws.cell(row=current_row, column=1, value=f"Valid only - Semester GPA: {valid_sem_gpa}, Cumulative GPA: {valid_cum_gpa}")
                current_row += 1
            
            current_row += 1  # Empty line between semesters
        
        # Detailed Results Sheet
        details_ws = wb.create_sheet("Detailed Results")
        
        # Headers
        headers = ["Semester", "Course Code", "Course Name", "Grade", "Credits", "Valid", "Issue Type", "Reason"]
        for col_num, header in enumerate(headers, 1):
            cell = details_ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Data rows
        row_num = 2
        for result in validation_results:
            is_valid = result.get("is_valid", True)
            
            details_ws.cell(row=row_num, column=1, value=result.get("semester", ""))
            details_ws.cell(row=row_num, column=2, value=result.get("course_code", ""))
            details_ws.cell(row=row_num, column=3, value=result.get("course_name", ""))
            details_ws.cell(row=row_num, column=4, value=result.get("grade", ""))
            
            # Get credits from semester data
            credits = ""
            semester_idx = result.get("semester_index", -1)
            if semester_idx >= 0 and semester_idx < len(semesters):
                for course in semesters[semester_idx].get("courses", []):
                    if course.get("code") == result.get("course_code"):
                        credits = course.get("credits", "")
                        break
            
            details_ws.cell(row=row_num, column=5, value=credits)
            details_ws.cell(row=row_num, column=6, value="Yes" if is_valid else "No")
            details_ws.cell(row=row_num, column=7, value=result.get("type", ""))
            details_ws.cell(row=row_num, column=8, value=result.get("reason", ""))
            
            # Color invalid courses red
            if not is_valid:
                for col in range(1, 9):
                    details_ws.cell(row=row_num, column=col).fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            
            row_num += 1
        
        # Auto-adjust column widths for both sheets
        for ws in [summary_ws, details_ws]:
            for column in ws.columns:
                max_length = 0
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
        wb.save(file_path)
        logger.info(f"Saved validation report to Excel: {file_path}")
        return True
        
    except ImportError:
        logger.error("openpyxl not installed. Cannot export to Excel.")
        return False
    except Exception as e:
        logger.error(f"Failed to save Excel file: {e}")
        return False

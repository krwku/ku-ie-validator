import streamlit as st
import json
import sys
from pathlib import Path
import tempfile
import os
import io
import PyPDF2

# Add your existing modules to path
sys.path.append(str(Path(__file__).parent))

from utils.pdf_extractor import PDFExtractor
from utils.validation_adapter import ValidationAdapter
from validator import CourseRegistrationValidator

def classify_course(course_code, course_name=""):
    """
    Classify course into appropriate category based on course code.
    Returns: (category, subcategory)
    """
    code = course_code.upper()
    
    # IE Core Courses (206xxx, 204111, 208xxx, 213xxx, 403xxx, 417xxx, 420xxx, 205xxx)
    if any(code.startswith(prefix) for prefix in ["01206", "01204111", "01208", "01213", "01403", "01417", "01420", "01205"]):
        return ("ie_core", "core")
    
    # Gen-Ed Classification
    if code.startswith("01175") or code.startswith("01101102") or code.startswith("01173151") or code.startswith("01999"):
        if any(keyword in course_name.lower() for keyword in ["health", "sport", "wellness", "aids", "food", "track", "tennis", "swim"]):
            return ("gen_ed", "wellness")
        elif any(keyword in course_name.lower() for keyword in ["art", "music", "design", "culture", "aesthetic"]):
            return ("gen_ed", "aesthetics")
    
    if any(code.startswith(prefix) for prefix in ["01131", "01132", "01005", "01200", "01418102"]):
        return ("gen_ed", "entrepreneurship")
    
    if any(code.startswith(prefix) for prefix in ["01361", "01355", "01354", "01356", "01357", "01358", "01362", "01367", "01395", "01398", "01399", "01371", "01418104", "01418111"]):
        return ("gen_ed", "language_communication")
    
    if any(code.startswith(prefix) for prefix in ["01001", "01015", "01174", "01350", "01387104", "01390", "01301", "01455", "01460"]) or "999111" in code:
        return ("gen_ed", "thai_citizen_global")
    
    if any(code.startswith(prefix) for prefix in ["01007", "01240", "01373", "01376", "01387102", "01420201"]):
        return ("gen_ed", "aesthetics")
    
    # Technical Electives (advanced 206xxx courses not in core)
    if code.startswith("01206") and any(advanced in code for advanced in ["411", "412", "413", "421", "422", "423", "441", "442", "443", "461", "462", "463", "464", "490", "496", "498"]):
        return ("technical_electives", "technical")
    
    # Free Electives (everything else)
    return ("free_electives", "free")

def create_smart_registration_excel(student_info, semesters, validation_results):
    """
    Create a smart, dynamic Excel registration format that:
    1. Properly classifies and places courses
    2. Uses dynamic row counts based on actual courses
    3. Has better formatting with separate code/name columns
    4. Removes unnecessary elements
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Registration Plan"
        
        # Define styles
        header_font = Font(bold=True, size=11)
        subheader_font = Font(bold=True, size=10)
        normal_font = Font(size=9)
        small_font = Font(size=8)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Color fills
        green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        gray_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        blue_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        red_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        
        # Set column widths for better readability
        column_widths = {
            'A': 12, 'B': 12, 'C': 35, 'D': 8, 'E': 8, 'F': 12, 'G': 35, 'H': 8, 'I': 8,
            'J': 12, 'K': 35, 'L': 8, 'M': 8, 'N': 12, 'O': 35, 'P': 8, 'Q': 8, 'R': 15, 'S': 10
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # HEADER SECTION - More informative
        ws['A1'] = "KU INDUSTRIAL ENGINEERING - COURSE REGISTRATION PLAN"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:Q1')
        ws['A1'].alignment = center_align
        ws['A1'].fill = blue_fill
        
        # Student Info
        ws['A3'] = "Student ID:"
        ws['B3'] = student_info.get('id', '')
        ws['B3'].fill = yellow_fill
        ws['E3'] = "Name:"
        ws['F3'] = student_info.get('name', '')
        ws['F3'].fill = yellow_fill
        
        ws['A4'] = "Field of Study:"
        ws['B4'] = student_info.get('field_of_study', 'Industrial Engineering')
        ws['E4'] = "Admission Date:"
        ws['F4'] = student_info.get('date_admission', '')
        
        # Summary Statistics
        total_credits = sum(
            course.get('credits', 0) 
            for semester in semesters 
            for course in semester.get('courses', [])
            if course.get('grade') not in ['F', 'W', 'N', '']
        )
        
        invalid_count = len([r for r in validation_results if not r.get("is_valid", True) and r.get("course_code") != "CREDIT_LIMIT"])
        
        ws['M3'] = "Total Credits:"
        ws['N3'] = total_credits
        ws['N3'].fill = green_fill if total_credits >= 130 else yellow_fill
        
        ws['M4'] = "Validation Issues:"
        ws['N4'] = invalid_count
        ws['N4'].fill = green_fill if invalid_count == 0 else red_fill
        
        # CLASSIFY ALL COURSES
        classified_courses = {
            "ie_core": [],
            "gen_ed": {
                "wellness": [],
                "entrepreneurship": [],
                "language_communication": [],
                "thai_citizen_global": [],
                "aesthetics": []
            },
            "technical_electives": [],
            "free_electives": []
        }
        
        # Process all courses from all semesters
        for sem_idx, semester in enumerate(semesters):
            semester_name = semester.get("semester", f"Semester {sem_idx + 1}")
            year = semester.get("year_int", 0)
            semester_type = semester.get("semester_type", "")
            
            for course in semester.get("courses", []):
                course_code = course.get("code", "")
                course_name = course.get("name", "")
                grade = course.get("grade", "")
                credits = course.get("credits", 0)
                
                # Check if course is valid
                is_valid = True
                issue_reason = ""
                for result in validation_results:
                    if (result.get('course_code') == course_code and 
                        not result.get('is_valid', True) and 
                        result.get('course_code') != 'CREDIT_LIMIT'):
                        is_valid = False
                        issue_reason = result.get('reason', '')
                        break
                
                # Classify course
                category, subcategory = classify_course(course_code, course_name)
                
                course_info = {
                    "code": course_code,
                    "name": course_name,
                    "grade": grade,
                    "credits": credits,
                    "semester": semester_name,
                    "is_valid": is_valid,
                    "issue": issue_reason,
                    "year": year,
                    "semester_type": semester_type
                }
                
                # Place in appropriate category
                if category == "ie_core":
                    classified_courses["ie_core"].append(course_info)
                elif category == "gen_ed":
                    classified_courses["gen_ed"][subcategory].append(course_info)
                elif category == "technical_electives":
                    classified_courses["technical_electives"].append(course_info)
                else:
                    classified_courses["free_electives"].append(course_info)
        
        # YEAR HEADERS - Simplified and cleaner
        current_row = 6
        
        # Year headers
        year_headers = ["Year 1", "Year 2", "Year 3", "Year 4"]
        year_cols = [['A', 'D'], ['E', 'H'], ['I', 'L'], ['M', 'P']]
        
        for i, (year_text, cols) in enumerate(zip(year_headers, year_cols)):
            start_col, end_col = cols[0], cols[1]
            ws.merge_cells(f'{start_col}{current_row}:{end_col}{current_row}')
            cell = ws[f'{start_col}{current_row}']
            cell.value = year_text
            cell.font = header_font
            cell.alignment = center_align
            cell.fill = gray_fill
            cell.border = border
        
        current_row += 1
        
        # Semester subheaders
        semester_headers = ["1st Sem", "2nd Sem"] * 4
        semester_cols = ['A', 'C', 'E', 'G', 'I', 'K', 'M', 'O']
        
        for i, (sem_text, col) in enumerate(zip(semester_headers, semester_cols)):
            end_col = chr(ord(col) + 1)
            ws.merge_cells(f'{col}{current_row}:{end_col}{current_row}')
            cell = ws[f'{col}{current_row}']
            cell.value = sem_text
            cell.font = subheader_font
            cell.alignment = center_align
            cell.border = border
        
        current_row += 1
        
        # Column headers for courses
        col_headers = ["Code", "Course Name", "Grade", "Credits"] * 4
        header_cols = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']
        
        for header, col in zip(col_headers, header_cols):
            cell = ws[f'{col}{current_row}']
            cell.value = header
            cell.font = subheader_font
            cell.alignment = center_align
            cell.fill = gray_fill
            cell.border = border
        
        current_row += 1
        
        # IE CORE COURSES SECTION
        ie_start_row = current_row
        
        # Section header
        ws[f'A{current_row}'] = "IE CORE COURSES"
        ws[f'A{current_row}'].font = header_font
        ws[f'A{current_row}'].fill = blue_fill
        ws.merge_cells(f'A{current_row}:P{current_row}')
        current_row += 1
        
        # Organize IE courses by year/semester
        ie_by_semester = {}
        for course in classified_courses["ie_core"]:
            year = course["year"]
            semester_type = course["semester_type"]
            if year > 0:
                min_year = min(s.get("year_int", 9999) for s in semesters if s.get("year_int", 0) > 0)
                grid_year = year - min_year + 1
                if grid_year <= 4:
                    key = f"Year{grid_year}_{semester_type}"
                    if key not in ie_by_semester:
                        ie_by_semester[key] = []
                    ie_by_semester[key].append(course)
        
        # Find max courses in any semester for row count
        max_ie_courses = max([len(courses) for courses in ie_by_semester.values()] + [1])
        
        # Fill IE courses
        semester_positions = {
            'Year1_First': ['A', 'B', 'C', 'D'],
            'Year1_Second': ['E', 'F', 'G', 'H'],
            'Year2_First': ['I', 'J', 'K', 'L'],
            'Year2_Second': ['M', 'N', 'O', 'P'],
            'Year3_First': ['A', 'B', 'C', 'D'],  # Will be in next section
            'Year3_Second': ['E', 'F', 'G', 'H'],
            'Year4_First': ['I', 'J', 'K', 'L'],
            'Year4_Second': ['M', 'N', 'O', 'P']
        }
        
        # Years 1-2
        for row_offset in range(max_ie_courses):
            row = current_row + row_offset
            
            for sem_key in ['Year1_First', 'Year1_Second', 'Year2_First', 'Year2_Second']:
                if sem_key in ie_by_semester and row_offset < len(ie_by_semester[sem_key]):
                    course = ie_by_semester[sem_key][row_offset]
                    cols = semester_positions[sem_key]
                    
                    # Course code
                    ws[f'{cols[0]}{row}'] = course["code"]
                    ws[f'{cols[0]}{row}'].border = border
                    ws[f'{cols[0]}{row}'].font = small_font
                    
                    # Course name
                    ws[f'{cols[1]}{row}'] = course["name"]
                    ws[f'{cols[1]}{row}'].border = border
                    ws[f'{cols[1]}{row}'].font = small_font
                    ws[f'{cols[1]}{row}'].alignment = left_align
                    
                    # Grade
                    ws[f'{cols[2]}{row}'] = course["grade"]
                    ws[f'{cols[2]}{row}'].border = border
                    ws[f'{cols[2]}{row}'].font = small_font
                    ws[f'{cols[2]}{row}'].alignment = center_align
                    
                    # Credits
                    ws[f'{cols[3]}{row}'] = course["credits"]
                    ws[f'{cols[3]}{row}'].border = border
                    ws[f'{cols[3]}{row}'].font = small_font
                    ws[f'{cols[3]}{row}'].alignment = center_align
                    
                    # Color coding
                    if not course["is_valid"]:
                        for col in cols:
                            ws[f'{col}{row}'].fill = red_fill
                    elif course["grade"] not in ['F', 'W', 'N', '']:
                        for col in cols:
                            ws[f'{col}{row}'].fill = green_fill
                else:
                    # Empty cells with borders
                    cols = semester_positions[sem_key]
                    for col in cols:
                        ws[f'{col}{row}'].border = border
        
        current_row += max_ie_courses + 2
        
        # Years 3-4 IE Courses
        if any(key.startswith(('Year3', 'Year4')) for key in ie_by_semester.keys()):
            ws[f'A{current_row}'] = "IE CORE COURSES (Years 3-4)"
            ws[f'A{current_row}'].font = header_font
            ws[f'A{current_row}'].fill = blue_fill
            ws.merge_cells(f'A{current_row}:P{current_row}')
            current_row += 1
            
            max_ie_courses_34 = max([len(ie_by_semester.get(key, [])) for key in ['Year3_First', 'Year3_Second', 'Year4_First', 'Year4_Second']] + [1])
            
            for row_offset in range(max_ie_courses_34):
                row = current_row + row_offset
                
                for sem_key in ['Year3_First', 'Year3_Second', 'Year4_First', 'Year4_Second']:
                    if sem_key in ie_by_semester and row_offset < len(ie_by_semester[sem_key]):
                        course = ie_by_semester[sem_key][row_offset]
                        cols = semester_positions[sem_key]
                        
                        # Fill course data (same as above)
                        ws[f'{cols[0]}{row}'] = course["code"]
                        ws[f'{cols[0]}{row}'].border = border
                        ws[f'{cols[0]}{row}'].font = small_font
                        
                        ws[f'{cols[1]}{row}'] = course["name"]
                        ws[f'{cols[1]}{row}'].border = border
                        ws[f'{cols[1]}{row}'].font = small_font
                        ws[f'{cols[1]}{row}'].alignment = left_align
                        
                        ws[f'{cols[2]}{row}'] = course["grade"]
                        ws[f'{cols[2]}{row}'].border = border
                        ws[f'{cols[2]}{row}'].font = small_font
                        ws[f'{cols[2]}{row}'].alignment = center_align
                        
                        ws[f'{cols[3]}{row}'] = course["credits"]
                        ws[f'{cols[3]}{row}'].border = border
                        ws[f'{cols[3]}{row}'].font = small_font
                        ws[f'{cols[3]}{row}'].alignment = center_align
                        
                        # Color coding
                        if not course["is_valid"]:
                            for col in cols:
                                ws[f'{col}{row}'].fill = red_fill
                        elif course["grade"] not in ['F', 'W', 'N', '']:
                            for col in cols:
                                ws[f'{col}{row}'].fill = green_fill
                    else:
                        # Empty cells with borders
                        cols = semester_positions[sem_key]
                        for col in cols:
                            ws[f'{col}{row}'].border = border
            
            current_row += max_ie_courses_34 + 2
        
        # GEN-ED SECTIONS - Dynamic based on actual courses
        gen_ed_categories = [
            ("wellness", "WELLNESS (7 Credits Required)", gray_fill),
            ("entrepreneurship", "ENTREPRENEURSHIP (5 Credits Required)", gray_fill),
            ("language_communication", "LANGUAGE & COMMUNICATION (15 Credits Required)", gray_fill),
            ("thai_citizen_global", "THAI CITIZEN & GLOBAL CITIZENSHIP (2 Credits Required)", gray_fill),
            ("aesthetics", "AESTHETICS (3 Credits Required)", gray_fill)
        ]
        
        for category_key, category_name, category_fill in gen_ed_categories:
            courses = classified_courses["gen_ed"][category_key]
            if courses:  # Only show if there are courses
                # Category header
                ws[f'A{current_row}'] = category_name
                ws[f'A{current_row}'].font = header_font
                ws[f'A{current_row}'].fill = category_fill
                ws.merge_cells(f'A{current_row}:P{current_row}')
                current_row += 1
                
                # Courses in this category
                for i, course in enumerate(courses):
                    row = current_row + i
                    
                    # Use first 4 columns for course info
                    ws[f'A{row}'] = course["code"]
                    ws[f'A{row}'].border = border
                    ws[f'A{row}'].font = small_font
                    
                    ws[f'B{row}'] = course["name"]
                    ws[f'B{row}'].border = border
                    ws[f'B{row}'].font = small_font
                    ws[f'B{row}'].alignment = left_align
                    
                    ws[f'C{row}'] = course["grade"]
                    ws[f'C{row}'].border = border
                    ws[f'C{row}'].font = small_font
                    ws[f'C{row}'].alignment = center_align
                    
                    ws[f'D{row}'] = course["credits"]
                    ws[f'D{row}'].border = border
                    ws[f'D{row}'].font = small_font
                    ws[f'D{row}'].alignment = center_align
                    
                    ws[f'E{row}'] = course["semester"]
                    ws[f'E{row}'].border = border
                    ws[f'E{row}'].font = small_font
                    
                    # Color coding
                    if not course["is_valid"]:
                        for col in ['A', 'B', 'C', 'D', 'E']:
                            ws[f'{col}{row}'].fill = red_fill
                    elif course["grade"] not in ['F', 'W', 'N', '']:
                        for col in ['A', 'B', 'C', 'D', 'E']:
                            ws[f'{col}{row}'].fill = green_fill
                
                current_row += len(courses) + 1
        
        # TECHNICAL ELECTIVES SECTION
        if classified_courses["technical_electives"]:
            ws[f'A{current_row}'] = "TECHNICAL ELECTIVES"
            ws[f'A{current_row}'].font = header_font
            ws[f'A{current_row}'].fill = blue_fill
            ws.merge_cells(f'A{current_row}:P{current_row}')
            current_row += 1
            
            for i, course in enumerate(classified_courses["technical_electives"]):
                row = current_row + i
                
                ws[f'A{row}'] = course["code"]
                ws[f'A{row}'].border = border
                ws[f'A{row}'].font = small_font
                
                ws[f'B{row}'] = course["name"]
                ws[f'B{row}'].border = border
                ws[f'B{row}'].font = small_font
                ws[f'B{row}'].alignment = left_align
                
                ws[f'C{row}'] = course["grade"]
                ws[f'C{row}'].border = border
                ws[f'C{row}'].font = small_font
                ws[f'C{row}'].alignment = center_align
                
                ws[f'D{row}'] = course["credits"]
                ws[f'D{row}'].border = border
                ws[f'D{row}'].font = small_font
                ws[f'D{row}'].alignment = center_align
                
                ws[f'E{row}'] = course["semester"]
                ws[f'E{row}'].border = border
                ws[f'E{row}'].font = small_font
                
                # Color coding
                if not course["is_valid"]:
                    for col in ['A', 'B', 'C', 'D', 'E']:
                        ws[f'{col}{row}'].fill = red_fill
                elif course["grade"] not in ['F', 'W', 'N', '']:
                    for col in ['A', 'B', 'C', 'D', 'E']:
                        ws[f'{col}{row}'].fill = green_fill
            
            current_row += len(classified_courses["technical_electives"]) + 1
        
        # FREE ELECTIVES SECTION
        if classified_courses["free_electives"]:
            ws[f'A{current_row}'] = "FREE ELECTIVES"
            ws[f'A{current_row}'].font = header_font
            ws[f'A{current_row}'].fill = yellow_fill
            ws.merge_cells(f'A{current_row}:P{current_row}')
            current_row += 1
            
            for i, course in enumerate(classified_courses["free_electives"]):
                row = current_row + i
                
                ws[f'A{row}'] = course["code"]
                ws[f'A{row}'].border = border
                ws[f'A{row}'].font = small_font
                
                ws[f'B{row}'] = course["name"]
                ws[f'B{row}'].border = border
                ws[f'B{row}'].font = small_font
                ws[f'B{row}'].alignment = left_align
                
                ws[f'C{row}'] = course["grade"]
                ws[f'C{row}'].border = border
                ws[f'C{row}'].font = small_font
                ws[f'C{row}'].alignment = center_align
                
                ws[f'D{row}'] = course["credits"]
                ws[f'D{row}'].border = border
                ws[f'D{row}'].font = small_font
                ws[f'D{row}'].alignment = center_align
                
                ws[f'E{row}'] = course["semester"]
                ws[f'E{row}'].border = border
                ws[f'E{row}'].font = small_font
                
                # Color coding
                if not course["is_valid"]:
                    for col in ['A', 'B', 'C', 'D', 'E']:
                        ws[f'{col}{row}'].fill = red_fill
                elif course["grade"] not in ['F', 'W', 'N', '']:
                    for col in ['A', 'B', 'C', 'D', 'E']:
                        ws[f'{col}{row}'].fill = green_fill
        
        # SUMMARY AT THE END
        current_row += len(classified_courses["free_electives"]) + 3
        
        ws[f'A{current_row}'] = "CREDIT SUMMARY"
        ws[f'A{current_row}'].font = header_font
        ws[f'A{current_row}'].fill = blue_fill
        ws.merge_cells(f'A{current_row}:E{current_row}')
        current_row += 1
        
        # Calculate credits by category
        ie_credits = sum(c["credits"] for c in classified_courses["ie_core"] if c["grade"] not in ['F', 'W', 'N', ''])
        gen_ed_credits = sum(
            sum(c["credits"] for c in courses if c["grade"] not in ['F', 'W', 'N', ''])
            for courses in classified_courses["gen_ed"].values()
        )
        tech_credits = sum(c["credits"] for c in classified_courses["technical_electives"] if c["grade"] not in ['F', 'W', 'N', ''])
        free_credits = sum(c["credits"] for c in classified_courses["free_electives"] if c["grade"] not in ['F', 'W', 'N', ''])
        
        summary_data = [
            ("IE Core Courses:", ie_credits),
            ("Gen-Ed Courses:", gen_ed_credits),
            ("Technical Electives:", tech_credits),
            ("Free Electives:", free_credits),
            ("TOTAL CREDITS:", ie_credits + gen_ed_credits + tech_credits + free_credits)
        ]
        
        for i, (label, credits) in enumerate(summary_data):
            row = current_row + i
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = subheader_font
            ws[f'B{row}'] = credits
            ws[f'B{row}'].font = subheader_font
            ws[f'B{row}'].fill = green_fill if credits > 0 else yellow_fill
            
            if label.startswith("TOTAL"):
                ws[f'A{row}'].fill = blue_fill
                ws[f'B{row}'].fill = blue_fill
        
        # Save to bytes
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            wb.save(tmp_file.name)
            
            with open(tmp_file.name, 'rb') as f:
                excel_bytes = f.read()
            
            os.unlink(tmp_file.name)
            return excel_bytes
            
    except Exception as e:
        st.error(f"Error creating Excel file: {e}")
        import traceback
        st.error(traceback.format_exc())
        return None

# [Rest of the Streamlit app remains the same but uses the new function]
def extract_text_from_pdf_bytes(pdf_bytes):
    """Extract text from PDF bytes using PyPDF2"""
    try:
        pdf_file = io.BytesIO(pdf_bytes)
        reader = PyPDF2.PdfReader(pdf_bytes)
        
        all_text = []
        for page in reader.pages:
            page_text = page.extract_text()
            if page_text and page_text.strip():
                all_text.append(page_text)
        
        return "\n".join(all_text)
    
    except Exception as e:
        st.error(f"Error extracting text from PDF: {e}")
        return ""

@st.cache_data
def load_comprehensive_course_data():
    """Load all course data including Gen-Ed and Technical Electives"""
    course_data_dir = Path(__file__).parent / "course_data"
    
    # Try to load from existing files
    available_files = {}
    
    if course_data_dir.exists():
        for json_file in course_data_dir.glob("*.json"):
            try:
                with open(json_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                
                file_name = json_file.stem
                if "2560" in file_name:
                    display_name = "Industrial Engineering 2560 (2017-2021)"
                elif "2565" in file_name:
                    display_name = "Industrial Engineering 2565 (2022-2026)"
                else:
                    display_name = file_name.replace("_", " ").title()
                
                available_files[display_name] = {
                    'data': data,
                    'filename': json_file.name,
                    'path': str(json_file)
                }
            except Exception as e:
                st.error(f"Error loading {json_file.name}: {e}")
    
    return available_files

def main():
    st.set_page_config(
        page_title="KU IE Course Validator", 
        page_icon="🎓",
        layout="wide"
    )
    
    st.title("🎓 KU Industrial Engineering Course Validator")
    st.markdown("*Smart Registration Planning with Dynamic Course Classification*")
    st.markdown("*Created for Raphin P.*")
    
    # Initialize session state
    if 'student_info' not in st.session_state:
        st.session_state.student_info = {}
    if 'semesters' not in st.session_state:
        st.session_state.semesters = []
    if 'validation_results' not in st.session_state:
        st.session_state.validation_results = []
    if 'selected_course_data' not in st.session_state:
        st.session_state.selected_course_data = None
    if 'processing_complete' not in st.session_state:
        st.session_state.processing_complete = False
    
    # Load course data
    available_course_data = load_comprehensive_course_data()
    
    # Sidebar
    with st.sidebar:
        st.header("⚙️ Configuration")
        
        if available_course_data:
            selected_catalog = st.selectbox(
                "📚 Select Course Catalog",
                options=list(available_course_data.keys()),
                help="Choose the course catalog for validation"
            )
            
            if selected_catalog:
                st.session_state.selected_course_data = available_course_data[selected_catalog]
                st.success(f"✅ Using: {available_course_data[selected_catalog]['filename']}")
        else:
            st.error("❌ No course data files found")
            st.stop()
        
        st.divider()
        
        st.header("📁 Upload Transcript")
        pdf_file = st.file_uploader(
            "Upload PDF Transcript", 
            type=['pdf'],
            help="Upload student transcript PDF"
        )
        
        if pdf_file is not None:
            st.info(f"📄 File: {pdf_file.name}")
            st.info(f"📊 Size: {len(pdf_file.getvalue()) / 1024:.1f} KB")
            
            if 'last_pdf_name' not in st.session_state or st.session_state.last_pdf_name != pdf_file.name:
                st.session_state.processing_complete = False
                st.session_state.student_info = {}
                st.session_state.semesters = []
                st.session_state.validation_results = []
                st.session_state.last_pdf_name = pdf_file.name
    
    # Main processing
    if pdf_file is not None and st.session_state.selected_course_data is not None:
        
        if not st.session_state.processing_complete:
            with st.spinner("🔄 Processing PDF and creating smart course classification..."):
                
                pdf_bytes = pdf_file.getvalue()
                extracted_text = extract_text_from_pdf_bytes(pdf_bytes)
                
                if not extracted_text:
                    st.error("❌ No text extracted from PDF")
                    st.stop()
                
                try:
                    extractor = PDFExtractor()
                    student_info, semesters, _ = extractor.process_pdf(None, extracted_text)
                    
                    if not student_info or not semesters:
                        st.error("❌ Failed to process transcript data")
                        st.stop()
                    
                    # Validate data
                    with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False) as tmp_file:
                        json.dump(st.session_state.selected_course_data['data'], tmp_file)
                        tmp_path = tmp_file.name
                    
                    validator = CourseRegistrationValidator(tmp_path)
                    passed_courses_history = validator.build_passed_courses_history(semesters)
                    
                    all_results = []
                    for i, semester in enumerate(semesters):
                        credit_valid, credit_reason = validator.validate_credit_limit(semester)
                        if not credit_valid:
                            all_results.append({
                                "semester": semester.get("semester", ""),
                                "semester_index": i,
                                "course_code": "CREDIT_LIMIT", 
                                "course_name": "Credit Limit Check",
                                "grade": "N/A",
                                "is_valid": True,
                                "reason": credit_reason,
                                "type": "credit_limit"
                            })
                        
                        for course in semester.get("courses", []):
                            is_valid, reason = validator.validate_course(
                                course, i, semesters, passed_courses_history, all_results
                            )
                            
                            all_results.append({
                                "semester": semester.get("semester", ""),
                                "semester_index": i,
                                "course_code": course.get("code", ""),
                                "course_name": course.get("name", ""),
                                "grade": course.get("grade", ""),
                                "is_valid": is_valid,
                                "reason": reason,
                                "type": "prerequisite"
                            })
                    
                    validator.propagate_invalidation(semesters, all_results)
                    os.unlink(tmp_path)
                    
                    st.session_state.student_info = student_info
                    st.session_state.semesters = semesters
                    st.session_state.validation_results = all_results
                    st.session_state.processing_complete = True
                    
                except Exception as e:
                    st.error(f"❌ Error during processing: {e}")
                    st.stop()
        
        # Display results
        if st.session_state.processing_complete:
            
            col1, col2 = st.columns([1, 1])
            
            with col1:
                st.header("📋 Student Information")
                st.write(f"**Student ID:** {st.session_state.student_info.get('id', 'Unknown')}")
                st.write(f"**Name:** {st.session_state.student_info.get('name', 'Unknown')}")
                st.write(f"**Field of Study:** {st.session_state.student_info.get('field_of_study', 'Unknown')}")
                
                st.divider()
                st.subheader("📚 Semester Summary")
                for i, sem in enumerate(st.session_state.semesters):
                    semester_name = sem.get('semester', f'Semester {i+1}')
                    course_count = len(sem.get('courses', []))
                    total_credits = sem.get('total_credits', 0)
                    st.write(f"• **{semester_name}:** {course_count} courses, {total_credits} credits")
            
            with col2:
                st.header("✅ Validation Results")
                
                invalid_results = [r for r in st.session_state.validation_results 
                                 if not r.get("is_valid", True) and r.get("course_code") != "CREDIT_LIMIT"]
                total_courses = len([r for r in st.session_state.validation_results 
                                   if r.get("course_code") != "CREDIT_LIMIT"])
                
                if len(invalid_results) == 0:
                    st.success(f"🎉 **Excellent!** All {total_courses} registrations are valid!")
                else:
                    st.error(f"⚠️ **Issues Found:** {len(invalid_results)} invalid registrations")
                
                if invalid_results:
                    with st.expander("❌ Invalid Registrations", expanded=True):
                        for result in invalid_results:
                            st.error(f"**{result.get('semester')}:** {result.get('course_code')} - {result.get('course_name')}")
                            st.write(f"   *Issue:* {result.get('reason')}")
            
            # Download section
            st.divider()
            st.header("📥 Download Smart Registration Reports")
            
            col_dl1, col_dl2, col_dl3 = st.columns(3)
            
            with col_dl1:
                # Generate Smart Excel format
                excel_bytes = create_smart_registration_excel(
                    st.session_state.student_info,
                    st.session_state.semesters,
                    st.session_state.validation_results
                )
                
                if excel_bytes:
                    st.download_button(
                        label="📋 Download Smart Registration Plan (.xlsx)",
                        data=excel_bytes,
                        file_name=f"KU_IE_smart_plan_{st.session_state.student_info.get('id', 'unknown')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        help="Intelligent course classification with dynamic layout and proper categorization",
                        use_container_width=True
                    )
                else:
                    st.error("❌ Failed to generate Excel file")
            
            with col_dl2:
                # Text report
                validator = CourseRegistrationValidator(
                    str(Path(__file__).parent / "course_data" / st.session_state.selected_course_data['filename'])
                )
                report_text = validator.generate_summary_report(
                    st.session_state.student_info, 
                    st.session_state.semesters, 
                    st.session_state.validation_results
                )
                
                st.download_button(
                    label="📄 Download Validation Report",
                    data=report_text,
                    file_name=f"validation_report_{st.session_state.student_info.get('id', 'unknown')}.txt",
                    mime="text/plain",
                    use_container_width=True
                )
            
            with col_dl3:
                # JSON data
                export_data = {
                    "student_info": st.session_state.student_info,
                    "semesters": st.session_state.semesters,
                    "validation_results": st.session_state.validation_results
                }
                
                st.download_button(
                    label="💾 Download Raw Data",
                    data=json.dumps(export_data, indent=2),
                    file_name=f"transcript_data_{st.session_state.student_info.get('id', 'unknown')}.json",
                    mime="application/json",
                    use_container_width=True
                )
            
            # Process another file
            st.divider()
            if st.button("🔄 Process Another PDF", type="secondary"):
                st.session_state.processing_complete = False
                st.session_state.student_info = {}
                st.session_state.semesters = []
                st.session_state.validation_results = []
                if 'last_pdf_name' in st.session_state:
                    del st.session_state.last_pdf_name
                st.rerun()
    
    else:
        # Welcome screen
        st.info("📋 **Ready for smart course validation and planning!**")
        
        col_info1, col_info2 = st.columns([1, 1])
        
        with col_info1:
            st.markdown("### 🎯 How to use:")
            st.markdown("""
            1. **Select course catalog** (IE 2560 or 2565)
            2. **Upload PDF transcript** in the sidebar
            3. **Wait for smart processing** ⚡
            4. **Download intelligent registration plan** 📋
            """)
        
        with col_info2:
            st.markdown("### 🧠 Smart Features:")
            st.markdown("• **Automatic course classification** - IE Core, Gen-Ed, Electives")
            st.markdown("• **Dynamic layout** - Adjusts to your actual courses")
            st.markdown("• **Clear formatting** - Separate columns for code/name")
            st.markdown("• **Color-coded validation** - Green/Red status")
            st.markdown("• **Credit summary** - Automatic calculations by category")
    
    # Status bar at bottom with your motto
    st.divider()
    col_status1, col_status2 = st.columns([3, 1])
    with col_status2:
        st.markdown("*Created for Raphin P.*", 
                   help="This application was specially created for Raphin P.")

if __name__ == "__main__":
    main()
